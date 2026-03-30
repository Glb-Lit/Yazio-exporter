from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.worksheet.worksheet import Worksheet


def _auto_width(sheet: Worksheet) -> None:
    for col in sheet.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value is None:
                continue
            max_len = max(max_len, len(str(cell.value)))
        sheet.column_dimensions[col_letter].width = min(max(max_len + 2, 12), 40)


def _style_header(sheet: Worksheet) -> None:
    for cell in sheet[1]:
        cell.font = Font(bold=True)
    sheet.freeze_panes = "A2"
    if sheet.max_column > 0 and sheet.max_row > 0:
        sheet.auto_filter.ref = sheet.dimensions
    _auto_width(sheet)


def build_excel_report(
    output_path: Path,
    nutrition_rows: list[dict[str, Any]],
    meal_rows: list[list[Any]],
    daily_rows: list[list[Any]],
) -> None:
    workbook = Workbook()

    nutrition_sheet = workbook.active
    nutrition_sheet.title = "nutrition_log"
    headers = [
        "Date",
        "Time",
        "Meal",
        "Product",
        "Amount",
        "Unit",
        "Portions",
        "Calories/g",
        "Calories total",
        "Protein/g",
        "Fat/g",
        "Carbs/g",
        "Protein total",
        "Fat total",
        "Carbs total",
    ]
    nutrition_sheet.append(headers)
    for row in nutrition_rows:
        nutrition_sheet.append([row.get(header, "") for header in headers])
    _style_header(nutrition_sheet)

    meal_sheet = workbook.create_sheet(title="meal_summary")
    for row in meal_rows:
        meal_sheet.append(row)
    _style_header(meal_sheet)

    daily_sheet = workbook.create_sheet(title="daily_summary")
    for row in daily_rows:
        daily_sheet.append(row)
    _style_header(daily_sheet)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
